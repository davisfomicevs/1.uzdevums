from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws = wb.active
total = 0
max_col = ws.max_column
max_row = ws.max_row

for kolonna in range(2,max_row + 1):
    id = ws['A'+str(kolonna)].value
    hours = ws['B' + str(kolonna)].value
    rate = ws['C'+str(kolonna)].value
    if (type(rate)!=str and type(hours)!=str):
        salary = rate*hours
        ws['D'+str(kolonna)].value=salary
        if salary > 3000:
            total += 1
print(total)

wb.save('result.xlsx')
wb.close()